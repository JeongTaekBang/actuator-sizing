"""
Noah Actuator Sizing Tool - Excel Workbook Generator
Creates the basic structure with sheets, data, and formatting
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

def create_workbook():
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create sheets in order
    ws_settings = wb.create_sheet("Settings")
    ws_valvelist = wb.create_sheet("ValveList")
    ws_config = wb.create_sheet("Configuration")
    # Normalized actuator DB structure
    ws_models = wb.create_sheet("DB_Models")
    ws_power = wb.create_sheet("DB_PowerOptions")
    ws_enclosure = wb.create_sheet("DB_EnclosureOptions")
    ws_electrical = wb.create_sheet("DB_ElectricalData")
    # Gearbox stays flat
    ws_gearboxes = wb.create_sheet("DB_Gearboxes")
    ws_couplings = wb.create_sheet("DB_Couplings")
    ws_options = wb.create_sheet("DB_Options")
    ws_datasheet = wb.create_sheet("Template_Datasheet")

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ==================== Settings Sheet ====================
    setup_settings_sheet(ws_settings, header_font, thin_border)

    # ==================== ValveList Sheet ====================
    setup_valvelist_sheet(ws_valvelist, header_font_white, header_fill, thin_border)

    # ==================== Configuration Sheet ====================
    setup_configuration_sheet(ws_config, header_font_white, header_fill, thin_border)

    # ==================== Normalized Actuator DB Sheets ====================
    setup_models_db(ws_models, header_font_white, header_fill, thin_border)
    setup_power_options_db(ws_power, header_font_white, header_fill, thin_border)
    setup_enclosure_options_db(ws_enclosure, header_font_white, header_fill, thin_border)
    setup_electrical_data_db(ws_electrical, header_font_white, header_fill, thin_border)

    # ==================== DB_Gearboxes Sheet ====================
    setup_gearboxes_db(ws_gearboxes, header_font_white, header_fill, thin_border)

    # ==================== DB_Couplings Sheet ====================
    setup_couplings_db(ws_couplings, header_font_white, header_fill, thin_border)

    # ==================== DB_Options Sheet ====================
    setup_options_db(ws_options, header_font_white, header_fill, thin_border)

    # ==================== Template_Datasheet Sheet ====================
    setup_datasheet_template(ws_datasheet, header_font, thin_border)

    # Hide DB sheets
    ws_models.sheet_state = 'hidden'
    ws_power.sheet_state = 'hidden'
    ws_enclosure.sheet_state = 'hidden'
    ws_electrical.sheet_state = 'hidden'
    ws_gearboxes.sheet_state = 'hidden'
    ws_couplings.sheet_state = 'hidden'
    ws_options.sheet_state = 'hidden'
    ws_datasheet.sheet_state = 'hidden'

    # Set active sheet to Settings
    wb.active = ws_settings

    return wb


def setup_settings_sheet(ws, header_font, border):
    """Setup Settings sheet with input fields and dropdowns"""

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15

    # Title
    ws['A1'] = "Noah Actuator Sizing Tool"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:C1')

    ws['A2'] = "Settings"
    ws['A2'].font = Font(bold=True, size=14)

    # Settings fields
    settings = [
        ("Torque Unit", "Nm", ["Nm", "lbf.ft", "kgf.m"]),
        ("Thrust Unit", "kN", ["kN", "lbf", "kgf"]),
        ("Enclosure", "Waterproof", ["Waterproof", "Explosionproof"]),
        ("Safety Factor", 1.25, None),
        ("Actuator Type", "Multi-turn", ["Multi-turn", "Part-turn", "Linear"]),
        ("Operation Mode", "On-Off", ["On-Off", "Modulating", "Modulating (High-Speed)"]),
        ("Fail-safe", "None", ["None", "Close-on-Fail (SR)"]),
        ("Duty Cycle", "Any", ["Any", "Intermittent (S2)", "Continuous (S4)"]),
        ("Voltage (V)", 380, [12, 24, 110, 120, 220, 230, 380, 440]),
        ("Phase", 3, [1, 3, "DC"]),
        ("Frequency (Hz)", 50, [50, 60]),
        ("Op. Time Min (%)", -50, None),
        ("Op. Time Max (%)", 50, None),
        ("Coupling Type", "Thrust Base - Threaded", ["Thrust Base - Threaded", "Standard (Part-turn)"]),
        ("Model Range", "All", ["All", "NA", "SA", "SR", "MA", "MS", "NL"]),
        ("Lines to Add", 10, None),
    ]

    row = 4
    for label, default, options in settings:
        ws.cell(row=row, column=1, value=label).font = header_font
        cell = ws.cell(row=row, column=2, value=default)
        cell.border = border

        if options:
            # Create dropdown
            options_str = ",".join(str(o) for o in options)
            dv = DataValidation(type="list", formula1=f'"{options_str}"', allow_blank=False)
            dv.error = "Please select from the list"
            dv.errorTitle = "Invalid Input"
            ws.add_data_validation(dv)
            dv.add(cell)

        row += 1

    # Settings Version (after all settings rows)
    ws['A21'] = "Settings Version:"
    ws['B21'] = "3.0"


def setup_valvelist_sheet(ws, header_font, header_fill, border):
    """Setup ValveList sheet with input columns and result columns
    
    Rows 1-2: Reserved for buttons
    Row 3: Header
    Row 4+: Data
    
    Color Scheme:
    - Input columns: Blue header (#4472C4), Light blue data (#D6E3F8)
    - Result columns: Green header (#548235), Light green data (#E2EFDA)
    """

    # Input columns (Line No. added as first column)
    # Lift (mm): valve stem travel distance
    # Pitch (mm): thread pitch, Turns = Lift / Pitch for Multi-turn
    input_headers = ["Line No.", "Tag", "ValveType", "Size", "Class", "Torque", "Thrust",
                     "CouplingType", "CouplingDim", "Lift(mm)", "Pitch(mm)", "Op.Time(sec)"]

    # Result columns (ActualSF, MaxStemDim, kW added)
    result_headers = ["Model", "Gearbox", "RPM", "Ratio", "OutputFlange", "CalcTorque",
                      "CalcThrust", "CalcOpTime", "ActualSF", "MaxStemDim", "kW", "Price", "Status"]

    all_headers = input_headers + result_headers
    input_count = len(input_headers)
    result_count = len(result_headers)

    # Color definitions
    # Input: Blue theme
    input_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    input_data_fill = PatternFill(start_color="D6E3F8", end_color="D6E3F8", fill_type="solid")

    # Result: Green theme
    result_header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    result_data_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # Set column widths
    # Cols: LineNo, Tag, ValveType, Size, Class, Torque, Thrust, CouplingType, CouplingDim, Lift, Pitch, OpTime,
    #       Model, Gearbox, RPM, Ratio, OutputFlange, CalcTorque, CalcThrust, CalcOpTime, ActualSF, MaxStemDim, kW, Price, Status
    widths = [8, 12, 12, 8, 8, 10, 10, 22, 12, 10, 10, 12, 15, 12, 8, 8, 12, 12, 12, 12, 10, 12, 8, 10, 15]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Set row heights for button area (rows 1-2)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30

    # Write headers on row 3 with color differentiation
    for col, header in enumerate(all_headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        
        # Apply different colors for input vs result columns
        if col <= input_count:
            cell.fill = input_header_fill
        else:
            cell.fill = result_header_fill
    
    # Apply subtle background color to data area (rows 4-100) for visual distinction
    for row in range(4, 101):
        for col in range(1, len(all_headers) + 1):
            cell = ws.cell(row=row, column=col)
            if col <= input_count:
                cell.fill = input_data_fill
            else:
                cell.fill = result_data_fill
    
    # Set Ratio column (column 16) to Text format to prevent "4:1" being interpreted as time
    for row in range(4, 101):
        ws.cell(row=row, column=16).number_format = '@'

    # Add color legend in button area (row 1)
    # Input legend (blue)
    ws['O1'] = "■ Input"
    ws['O1'].font = Font(bold=True, size=9, color="4472C4")
    
    # Result legend (green)
    ws['P1'] = "■ Result"
    ws['P1'].font = Font(bold=True, size=9, color="548235")

    # Add CouplingType dropdown (Multi-turn + Part-turn options) - Column H
    coupling_types = "Thrust Base - Threaded,Standard (Part-turn)"
    dv_coupling = DataValidation(type="list", formula1=f'"{coupling_types}"', allow_blank=True)
    ws.add_data_validation(dv_coupling)
    dv_coupling.add(f'H4:H100')

    # Add ValveType dropdown - Column C
    valve_types = "Gate,Globe,Ball,Butterfly,Plug,Linear"
    dv_valve = DataValidation(type="list", formula1=f'"{valve_types}"', allow_blank=True)
    ws.add_data_validation(dv_valve)
    dv_valve.add(f'C4:C100')

    # Freeze rows 1-3 (button area + header)
    ws.freeze_panes = 'A4'


def setup_configuration_sheet(ws, header_font, header_fill, border):
    """Setup Configuration sheet for options and pricing
    
    Structure:
    Line | Tag | Model | Gearbox | Base | HTR | MOD | POS | LMT | EXD | Painting | Qty | Unit | Total
    """
    
    # Column widths
    widths = {
        'A': 8,   # Line
        'B': 12,  # Tag
        'C': 15,  # Model
        'D': 12,  # Gearbox
        'E': 12,  # Base$
        'F': 8,   # HTR
        'G': 8,   # MOD
        'H': 8,   # POS
        'I': 8,   # LMT
        'J': 8,   # EXD
        'K': 12,  # Painting
        'L': 8,   # Qty
        'M': 12,  # Unit$
        'N': 12,  # Total$
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Headers
    headers = [
        "Line", "Tag", "Model", "Gearbox", "Base",
        "HTR", "MOD", "POS", "LMT", "EXD",
        "Painting", "Qty", "Unit", "Total"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Yes/No dropdown for option columns (F-J)
    yesno_list = "Yes,No"
    dv_yesno = DataValidation(type="list", formula1=f'"{yesno_list}"', allow_blank=False)
    dv_yesno.error = "Please select Yes or No"
    dv_yesno.errorTitle = "Invalid Input"
    ws.add_data_validation(dv_yesno)
    dv_yesno.add('F2:J100')

    # Painting dropdown (K)
    paint_list = "None,PAINT-EP,PAINT-PU,PAINT-SPEC"
    dv_paint = DataValidation(type="list", formula1=f'"{paint_list}"', allow_blank=False)
    dv_paint.error = "Please select from the list"
    dv_paint.errorTitle = "Invalid Input"
    ws.add_data_validation(dv_paint)
    dv_paint.add('K2:K100')

    # Add border to data area
    for row in range(2, 52):
        for col in range(1, 15):
            cell = ws.cell(row=row, column=col)
            cell.border = border

    # Grand Total row
    ws['A55'] = "Grand Total:"
    ws['A55'].font = Font(bold=True)
    ws['N55'] = "=SUM(N2:N54)"
    ws['N55'].font = Font(bold=True)

    # Option price reference (hidden area for documentation)
    ws['A57'] = "Option Prices (Reference):"
    ws['A57'].font = Font(bold=True, size=10)
    option_prices = [
        ("HTR", "Space Heater", 50),
        ("MOD", "Modulating Control", 200),
        ("POS", "Position Transmitter", 150),
        ("LMT", "Limit Switch", 80),
        ("EXD", "Explosionproof Upgrade", 300),
    ]
    for i, (code, desc, price) in enumerate(option_prices, 58):
        ws.cell(row=i, column=1, value=code)
        ws.cell(row=i, column=2, value=desc)
        ws.cell(row=i, column=3, value=f"${price}")

    ws.freeze_panes = 'A2'


def setup_models_db(ws, header_font, header_fill, border):
    """Setup DB_Models sheet with flat structure for Noah actuators

    Each Model × kW × Hz × RPM (or Model × Type × Hz) combination is a separate row.
    This flat structure simplifies VBA lookup logic.

    18 columns:
    - Columns 1-16: Common fields (Multi-turn, Part-turn, Linear)
    - Columns 17-18: Linear-specific (Speed_mm_sec, Stroke_mm)

    Phase column (column 6):
    - MS series: 1 or 3 (1-phase vs 3-phase have different torque)
    - Other series: 0 (phase doesn't affect torque, handled by PowerOptions)
    """
    headers = ["Model", "Series", "ActType", "MotorPower_kW", "ControlType", "Phase",
               "Freq", "RPM", "Torque_Nm", "Thrust_kN", "OpTime_sec",
               "DutyCycle", "OutputFlange", "MaxStemDim_mm", "Weight_kg", "BasePrice",
               "Speed_mm_sec", "Stroke_mm"]

    data = []

    # ==================== NA Series (Part-turn) ====================
    # Simple structure: Model + Freq → fixed specs
    # From docs/noah_torque_tables.md NA specs table
    na_models = [
        # Model, Torque_Nm, OpTime_50Hz, OpTime_60Hz, MaxStemDim, Weight, DutyCycle%
        ("NA006", 60, 18, 16, 22, 11, 50),
        ("NA009", 90, 20, 17, 22, 11, 50),
        ("NA015", 150, 23, 20, 22, 14, 50),
        ("NA019", 190, 23, 20, 22, 14, 50),
        ("NA028", 280, 29, 25, 32, 17, 50),
        ("NA038", 380, 29, 25, 32, 17, 30),
        ("NA050", 500, 29, 25, 32, 17, 25),
        ("NA060", 600, 38, 31, 42, 24, 25),
        ("NA080", 800, 38, 31, 42, 25, 25),
        ("NA100", 1000, 38, 31, 42, 25, 25),
        ("NA120", 1200, 36, 30, 42, 25, 25),
        ("NA150", 1500, 114, 93, 75, 65, 25),
        ("NA200", 2000, 114, 93, 75, 65, 25),
        ("NA250", 2500, 114, 93, 75, 65, 25),
        ("NA300", 3000, 144, 120, 75, 65, 25),
        ("NA350", 3500, 144, 120, 75, 65, 25),
    ]

    # Output flange based on torque (approximate mapping)
    def get_na_flange(torque):
        if torque <= 280: return "F07"
        elif torque <= 800: return "F10"
        elif torque <= 1200: return "F14"
        else: return "F16"

    # Base price estimate based on torque
    def get_na_price(torque):
        if torque <= 100: return 500
        elif torque <= 300: return 700
        elif torque <= 600: return 900
        elif torque <= 1000: return 1200
        elif torque <= 2000: return 1600
        else: return 2000

    for model, torque, op50, op60, stem, weight, duty in na_models:
        # 50Hz version (Phase=0: phase doesn't affect NA torque)
        data.append([
            model, "NA", "Part-turn", None, None, 0,
            50, None, torque, None, op50,
            f"S4-{duty}%", get_na_flange(torque), stem, weight, get_na_price(torque),
            None, None  # Speed_mm_sec, Stroke_mm (not used for Part-turn)
        ])
        # 60Hz version
        data.append([
            model, "NA", "Part-turn", None, None, 0,
            60, None, torque, None, op60,
            f"S4-{duty}%", get_na_flange(torque), stem, weight, get_na_price(torque),
            None, None  # Speed_mm_sec, Stroke_mm (not used for Part-turn)
        ])

    # ==================== SA Series (Part-turn, small) ====================
    # Structure: Model × ControlType × Freq
    # From docs/noah_torque_tables.md SA specs table
    #
    # ControlType:
    #   - ONOFF: 기본 개폐제어
    #   - PCU: 비례제어 (Proportional Control Unit) - 표준 속도
    #   - SCP: 고속 비례제어 (Stepping Control Panel) - 동작시간 절반
    #
    # 방폭형 모델 (SA05X, SA09X):
    #   - ControlType은 ONOFF (개폐제어)
    #   - Enclosure = Exd (방폭형) - DB_EnclosureOptions에서 처리
    #   - 모델명에 'X' 접미사로 구분
    sa_models = [
        # Model, ControlType, Torque_Nm, OpTime_50Hz, OpTime_60Hz, Weight, MotorW
        ("SA003", "ONOFF", 30, 17, 15, 1.7, 25),
        ("SA005", "ONOFF", 50, 17, 14, 2.8, 6),
        ("SA005L", "PCU", 50, 17, 14, 3.2, 6),
        ("SA005L", "SCP", 50, 8, 8, 4.3, 15),
        ("SA05X", "ONOFF", 50, 17, 14, 5.0, 6),   # 방폭형 - Enclosure=Exd
        ("SA009", "ONOFF", 90, 32, 26, 2.8, 6),
        ("SA009L", "PCU", 90, 32, 26, 3.2, 6),
        ("SA009L", "SCP", 90, 13, 13, 4.3, 15),
        ("SA09X", "ONOFF", 90, 32, 26, 5.0, 6),   # 방폭형 - Enclosure=Exd
    ]

    def get_sa_flange(torque):
        if torque <= 30: return "F05"
        elif torque <= 50: return "F07"
        else: return "F10"

    def get_sa_price(model, ctrl_type, torque):
        base = 300 if torque <= 30 else (400 if torque <= 50 else 500)
        if ctrl_type == "SCP": base += 150   # 고속 제어판 추가
        if ctrl_type == "PCU": base += 50    # 비례제어 유닛 추가
        if "X" in model: base += 200         # 방폭형 추가 (Enclosure 비용)
        return base

    for model, ctrl_type, torque, op50, op60, weight, motor_w in sa_models:
        # 50Hz version (Phase=0: SA is 1-phase only, phase doesn't vary)
        data.append([
            model, "SA", "Part-turn", None, ctrl_type, 0,
            50, None, torque, None, op50,
            "S2-15min", get_sa_flange(torque), 20, weight, get_sa_price(model, ctrl_type, torque),
            None, None  # Speed_mm_sec, Stroke_mm (not used for Part-turn)
        ])
        # 60Hz version
        data.append([
            model, "SA", "Part-turn", None, ctrl_type, 0,
            60, None, torque, None, op60,
            "S2-15min", get_sa_flange(torque), 20, weight, get_sa_price(model, ctrl_type, torque),
            None, None  # Speed_mm_sec, Stroke_mm (not used for Part-turn)
        ])

    # ==================== MS Series (Multi-turn, small) ====================
    # Structure: Model × Phase × Freq × RPM
    # From docs/noah_torque_tables.md MS specs table
    #
    # IMPORTANT: MS series has different torque based on Phase:
    # - 3-phase: 110 Nm (high torque)
    # - 1-phase: 45 Nm (low torque)
    # Phase column distinguishes these (consistent with other series naming)

    # 3-phase / S2-30min (Phase=3)
    ms_3phase = [
        # Hz, RPM, Torque
        (50, 20.5, 110), (60, 24.5, 110),
        (50, 31.0, 110), (60, 37.0, 110),
        (50, 38.0, 100), (60, 46.0, 100),
        (50, 52.5, 100), (60, 63.5, 100),
        (50, 66.0, 80), (60, 79.5, 80),
        (50, 83.5, 75), (60, 101.0, 75),
    ]

    for hz, rpm, torque in ms_3phase:
        data.append([
            "MS01", "MS", "Multi-turn", None, None, 3,  # Phase=3
            hz, rpm, torque, 30, None,  # Thrust 30kN assumed
            "S2-30min", "F07", 40, 15, 600,
            None, None  # Speed_mm_sec, Stroke_mm (not used for Multi-turn)
        ])

    # 1-phase / S2-15min (Phase=1)
    ms_1phase = [
        (50, 19.0, 45), (60, 23.4, 45),
        (50, 28.7, 45), (60, 35.3, 45),
    ]

    for hz, rpm, torque in ms_1phase:
        data.append([
            "MS01", "MS", "Multi-turn", None, None, 1,  # Phase=1
            hz, rpm, torque, 20, None,  # Thrust 20kN assumed
            "S2-15min", "F07", 40, 15, 500,
            None, None  # Speed_mm_sec, Stroke_mm (not used for Multi-turn)
        ])

    # ==================== MA Series (Multi-turn) ====================
    # Structure: Model × kW × Freq × RPM
    # From docs/noah_torque_tables.md MA torque table
    # This creates many rows, so we include representative combinations

    # RPM values for 50Hz and 60Hz
    rpm_50hz = [16.0, 20.0, 27.7, 35.5, 47.7, 59.8, 82.5, 105.5, 120.5, 165.0]
    rpm_60hz = [19.2, 24.0, 33.3, 42.5, 57.1, 72.0, 98.8, 126.5, 146.0, 200.0]

    # MA01 data: kW, [(rpm_idx, torque_50hz), ...]
    ma01_data = [
        (0.2, "MA01", [(0, 88), (2, 88), (4, 88), (6, 80), (8, 59), (10, 49)]),  # Subset of RPM indices
        (0.4, "MA01", [(0, 138), (2, 138), (4, 121), (6, 93), (8, 75), (10, 55), (12, 40)]),
        (0.75, "MA01", [(0, 138), (2, 138), (4, 138), (6, 138), (8, 123), (10, 95), (12, 69), (14, 55), (16, 49)]),
    ]

    # MA02 data
    ma02_data = [
        (1.5, "MA02", [(0, 415), (2, 415), (4, 415), (6, 324), (8, 238), (10, 192), (12, 139), (14, 108), (16, 96), (18, 69)]),
        (2.2, "MA02", [(0, 415), (2, 415), (4, 415), (6, 415), (8, 344), (10, 278), (12, 202), (14, 156), (16, 139), (18, 101)]),
    ]

    # MA03 data
    ma03_data = [
        (2.2, "MA03", [(0, 845), (2, 845), (4, 605), (6, 470), (8, 345), (10, 280), (12, 200), (14, 155)]),
        (3.7, "MA03", [(0, 845), (2, 845), (4, 845), (6, 800), (8, 590), (10, 475), (12, 345), (14, 270), (16, 240), (18, 170)]),
        (5.5, "MA03", [(0, 845), (2, 845), (4, 845), (6, 845), (8, 845), (10, 700), (12, 501), (14, 390), (16, 345), (18, 250)]),
        (7.5, "MA03", [(6, 845), (8, 845), (10, 845), (12, 690), (14, 535), (16, 470), (18, 345)]),
    ]

    # MA04 data
    ma04_data = [
        (5.5, "MA04", [(0, 2800), (2, 2215), (4, 1620), (6, 1275), (8, 955), (10, 795), (12, 575), (14, 445)]),
        (7.5, "MA04", [(0, 3865), (2, 3040), (4, 2225), (6, 1750), (8, 1310), (10, 1090), (12, 785), (14, 610)]),
        (11, "MA04", [(0, 3920), (2, 3920), (4, 3230), (6, 2540), (8, 1905), (10, 1585), (12, 1145), (14, 885), (16, 790), (18, 570)]),
        (15, "MA04", [(0, 3920), (2, 3920), (4, 3920), (6, 3455), (8, 2595), (10, 2155), (12, 1555), (14, 1200), (16, 1065), (18, 770)]),
    ]

    # MA05 data
    ma05_data = [
        (11, "MA05", [(0, 5615), (2, 4415), (4, 3230), (6, 2540), (8, 1905), (10, 1585), (12, 1145), (14, 885)]),
        (15, "MA05", [(0, 7840), (2, 6010), (4, 4395), (6, 3455), (8, 2595), (10, 2155), (12, 1555), (14, 1200)]),
        (18.5, "MA05", [(0, 7840), (2, 7385), (4, 5400), (6, 4250), (8, 3185), (10, 2645), (12, 1910), (14, 1475), (16, 1325), (18, 955)]),
        (22, "MA05", [(0, 7840), (2, 7840), (4, 6405), (6, 5035), (8, 3780), (10, 3140), (12, 2265), (14, 1750), (16, 1555), (18, 1125)]),
    ]

    # MA06 data
    ma06_data = [
        (18.5, "MA06", [(0, 9390), (2, 7385), (4, 5400), (6, 4250), (8, 3185), (10, 2645), (12, 1910), (14, 1475)]),
        (22, "MA06", [(0, 11140), (2, 8760), (4, 6405), (6, 5035), (8, 3780), (10, 3140), (12, 2265), (14, 1750)]),
        (30, "MA06", [(0, 15680), (2, 12020), (4, 8790), (6, 6910), (8, 5185), (10, 4305), (12, 3105), (14, 2400), (16, 2125), (18, 1535)]),
        (37, "MA06", [(0, 15680), (2, 14770), (4, 10800), (6, 8495), (8, 6370), (10, 5290), (12, 3815), (14, 2950), (16, 2620), (18, 1890)]),
    ]

    def get_ma_flange(model):
        if model == "MA01": return "F10"
        elif model == "MA02": return "F10"
        elif model == "MA03": return "F14"
        elif model == "MA04": return "F16"
        elif model == "MA05": return "F25"
        else: return "F25"  # MA06

    def get_ma_stem(model):
        if model == "MA01": return 65
        elif model == "MA02": return 65
        elif model == "MA03": return 90
        elif model == "MA04": return 120
        elif model == "MA05": return 160
        else: return 160  # MA06

    def get_ma_weight(model, kw):
        base = {"MA01": 45, "MA02": 65, "MA03": 95, "MA04": 150, "MA05": 200, "MA06": 280}
        return base.get(model, 100) + int(kw * 5)

    def get_ma_thrust(model):
        # Approximate thrust in kN based on model
        thrust = {"MA01": 50, "MA02": 80, "MA03": 120, "MA04": 180, "MA05": 250, "MA06": 350}
        return thrust.get(model, 100)

    def get_ma_price(model, kw):
        base = {"MA01": 800, "MA02": 1200, "MA03": 1800, "MA04": 2800, "MA05": 4000, "MA06": 5500}
        return base.get(model, 1000) + int(kw * 100)

    all_ma_data = ma01_data + ma02_data + ma03_data + ma04_data + ma05_data + ma06_data

    for kw, model, rpm_torques in all_ma_data:
        for rpm_idx, torque in rpm_torques:
            # 50Hz version (Phase=0: MA is 3-phase only, phase doesn't vary)
            if rpm_idx < len(rpm_50hz):
                rpm_50 = rpm_50hz[rpm_idx // 2]  # Map index to actual RPM
                data.append([
                    model, "MA", "Multi-turn", kw, None, 0,
                    50, rpm_50, torque, get_ma_thrust(model), None,
                    "S2-30min", get_ma_flange(model), get_ma_stem(model),
                    get_ma_weight(model, kw), get_ma_price(model, kw),
                    None, None  # Speed_mm_sec, Stroke_mm (not used for Multi-turn)
                ])
            # 60Hz version (slightly different RPM)
            if rpm_idx < len(rpm_60hz):
                rpm_60 = rpm_60hz[rpm_idx // 2]
                data.append([
                    model, "MA", "Multi-turn", kw, None, 0,
                    60, rpm_60, torque, get_ma_thrust(model), None,
                    "S2-30min", get_ma_flange(model), get_ma_stem(model),
                    get_ma_weight(model, kw), get_ma_price(model, kw),
                    None, None  # Speed_mm_sec, Stroke_mm (not used for Multi-turn)
                ])

    # ==================== SR Series (Part-turn, Spring Return) ====================
    # Structure: Model × Freq
    # Spring Return actuators - motor driven Open, spring driven Close (1-2 sec)
    # OpTime = Open time (motor driven) for sizing
    # From user-provided SR specification table

    # SR models: Model, Torque_Nm, OpTime_50Hz (220VAC Open), OpTime_60Hz (220VAC Open), MaxStemDim, Weight, Flange
    sr_models = [
        ("SR05", 50, 17, 14, 20, 26.5, "F07"),
        ("SR10", 100, 20, 17, 22, 35, "F07"),
        ("SR20", 200, 59, 50, 32, 51, "F10"),
        ("SR30", 300, 87, 75, 42, 62, "F10"),
        ("SR50", 500, 116, 99, 42, 82, "F10"),
    ]

    def get_sr_price(torque):
        if torque <= 50: return 600
        elif torque <= 100: return 700
        elif torque <= 200: return 900
        elif torque <= 300: return 1100
        else: return 1200

    for model, torque, op50, op60, stem, weight, flange in sr_models:
        # 50Hz version (Phase=0: phase doesn't affect SR torque)
        data.append([
            model, "SR", "Part-turn", None, "SR", 0,
            50, None, torque, None, op50,
            "S2-15min", flange, stem, weight, get_sr_price(torque),
            None, None  # Speed_mm_sec, Stroke_mm (not used for Part-turn)
        ])
        # 60Hz version
        data.append([
            model, "SR", "Part-turn", None, "SR", 0,
            60, None, torque, None, op60,
            "S2-15min", flange, stem, weight, get_sr_price(torque),
            None, None  # Speed_mm_sec, Stroke_mm (not used for Part-turn)
        ])

    # ==================== NL Series (Linear) ====================
    # Structure: Model × Freq
    # Linear actuators - use Thrust (kN) instead of Torque
    # OpTime = Stroke / Speed
    # From user-provided NL specification table

    # NL models: Model, Thrust_kN, Speed_50Hz, Speed_60Hz, Stroke, Duty%, MaxStem(from thread), Weight, Motor_W
    nl_models = [
        ("NL04", 4, 0.8, 0.93, 40, 50, 20, 16, 15),
        ("NL06", 6, 0.79, 0.9, 40, 50, 20, 16, 25),
        ("NL08", 8, 0.75, 0.86, 50, 50, 20, 18, 25),
        ("NL10", 10, 0.72, 0.83, 50, 50, 20, 18, 40),
        ("NL20", 20, 0.85, 1.0, 100, 30, 24, 31, 60),
        ("NL25", 25, 0.72, 0.87, 100, 30, 24, 31, 90),
        ("NL35", 35, 0.4, 0.47, 100, 20, 24, 31, 90),
    ]

    def get_nl_price(thrust):
        if thrust <= 6: return 500
        elif thrust <= 10: return 700
        elif thrust <= 25: return 1000
        else: return 1200

    for model, thrust, spd50, spd60, stroke, duty, stem, weight, motor_w in nl_models:
        # 50Hz version (Phase=0: phase doesn't affect NL thrust)
        data.append([
            model, "NL", "Linear", None, None, 0,
            50, None, None, thrust, None,  # Torque=None, Thrust=thrust, OpTime=None (calculated)
            f"S4-{duty}%", None, stem, weight, get_nl_price(thrust),
            spd50, stroke  # Speed_mm_sec, Stroke_mm
        ])
        # 60Hz version
        data.append([
            model, "NL", "Linear", None, None, 0,
            60, None, None, thrust, None,
            f"S4-{duty}%", None, stem, weight, get_nl_price(thrust),
            spd60, stroke
        ])

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    # Write data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    # 18 columns width
    widths = [20, 8, 12, 14, 12, 8, 8, 8, 12, 12, 12, 12, 12, 14, 12, 12, 12, 10]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = 'A2'


def setup_power_options_db(ws, header_font, header_fill, border):
    """Setup DB_PowerOptions sheet (normalized)

    Each model can have multiple power configurations.
    PriceAdder is added to BasePrice from DB_Models.
    Model names must match DB_Models exactly.
    """
    headers = ["Model", "Voltage", "Phase", "Freq", "PriceAdder"]

    data = []

    # ==================== NA Series Power Options ====================
    # From docs/noah_torque_tables.md - NA supports DC and AC
    # DC 12V: NA006-NA060, DC 24V: NA006-NA080, AC all: NA006-NA350
    na_models_dc12 = ["NA006", "NA009", "NA015", "NA019", "NA028", "NA038", "NA050", "NA060"]
    na_models_dc24 = ["NA006", "NA009", "NA015", "NA019", "NA028", "NA038", "NA050", "NA060", "NA080"]
    na_all = ["NA006", "NA009", "NA015", "NA019", "NA028", "NA038", "NA050",
              "NA060", "NA080", "NA100", "NA120", "NA150", "NA200", "NA250", "NA300", "NA350"]

    for model in na_all:
        for hz in [50, 60]:
            # AC options (all models)
            data.append([model, 380, 3, hz, 0])      # 3-phase 380V (standard)
            data.append([model, 220, 1, hz, 0])      # 1-phase 220V
            data.append([model, 110, 1, hz, 0])      # 1-phase 110V
            data.append([model, 440, 3, hz, 50])     # 3-phase 440V
            # DC options (smaller models only)
            if model in na_models_dc24:
                data.append([model, 24, "DC", hz, 50])
            if model in na_models_dc12:
                data.append([model, 12, "DC", hz, 50])

    # ==================== SA Series Power Options ====================
    # SA is 1-phase only (380V/440V not supported)
    # ControlType: ONOFF, PCU, SCP (EXP는 Enclosure 옵션)
    sa_models = [
        ("SA003", ["ONOFF"]),
        ("SA005", ["ONOFF"]),
        ("SA005L", ["PCU", "SCP"]),
        ("SA05X", ["ONOFF"]),     # 방폭형 - ControlType은 ONOFF
        ("SA009", ["ONOFF"]),
        ("SA009L", ["PCU", "SCP"]),
        ("SA09X", ["ONOFF"]),     # 방폭형 - ControlType은 ONOFF
    ]

    for model, types in sa_models:
        for ctrl_type in types:
            for hz in [50, 60]:
                # 1-phase AC only
                data.append([model, 220, 1, hz, 0])   # 1-phase 220V (standard)
                data.append([model, 110, 1, hz, 0])   # 1-phase 110V
                data.append([model, 24, 1, hz, 0])    # 1-phase 24V AC
                # DC 24V option
                data.append([model, 24, "DC", hz, 50])

    # ==================== MS Series Power Options ====================
    # MS supports 1-phase and 3-phase (same Model "MS01", Phase column in DB_Models)
    # PowerOptions uses Model + Phase to match the correct DB_Models row

    # 3-phase power options (for MS01 Phase=3 rows in DB_Models)
    for hz in [50, 60]:
        data.append(["MS01", 380, 3, hz, 0])
        data.append(["MS01", 220, 3, hz, 0])
        data.append(["MS01", 440, 3, hz, 50])

    # 1-phase power options (for MS01 Phase=1 rows in DB_Models)
    for hz in [50, 60]:
        data.append(["MS01", 220, 1, hz, 0])
        data.append(["MS01", 110, 1, hz, 0])

    # ==================== MA Series Power Options ====================
    # MA is 3-phase only
    rpm_50hz = [16.0, 20.0, 27.7, 35.5, 47.7, 59.8, 82.5, 105.5, 120.5, 165.0]
    rpm_60hz = [19.2, 24.0, 33.3, 42.5, 57.1, 72.0, 98.8, 126.5, 146.0, 200.0]

    ma_configs = [
        ("MA01", [0.2, 0.4, 0.75]),
        ("MA02", [1.5, 2.2]),
        ("MA03", [2.2, 3.7, 5.5, 7.5]),
        ("MA04", [5.5, 7.5, 11, 15]),
        ("MA05", [11, 15, 18.5, 22]),
        ("MA06", [18.5, 22, 30, 37]),
    ]

    # Generate power options for MA models (simplified - same options for all variants)
    for model, kw_list in ma_configs:
        # 50Hz options
        data.append([model, 380, 3, 50, 0])
        data.append([model, 220, 3, 50, 0])
        data.append([model, 440, 3, 50, 50])
        # 60Hz options
        data.append([model, 380, 3, 60, 0])
        data.append([model, 440, 3, 60, 0])

    # ==================== SR Series Power Options ====================
    # SR supports: 110V 1P, 120V 1P, 220V 1P, 230V 1P, 380V 3P, 440V 3P, DC 24V
    sr_models_list = ["SR05", "SR10", "SR20", "SR30", "SR50"]

    for model in sr_models_list:
        for hz in [50, 60]:
            # AC options
            data.append([model, 110, 1, hz, 0])      # 1-phase 110V
            data.append([model, 120, 1, hz, 0])      # 1-phase 120V
            data.append([model, 220, 1, hz, 0])      # 1-phase 220V
            data.append([model, 230, 1, hz, 0])      # 1-phase 230V
            data.append([model, 380, 3, hz, 0])      # 3-phase 380V
            data.append([model, 440, 3, hz, 50])     # 3-phase 440V
            # DC option
            data.append([model, 24, "DC", hz, 50])   # DC 24V

    # ==================== NL Series Power Options ====================
    # NL supports: 110V 1P, 220V 1P, 380V 3P, 440V 3P, DC 24V
    nl_models_list = ["NL04", "NL06", "NL08", "NL10", "NL20", "NL25", "NL35"]

    for model in nl_models_list:
        for hz in [50, 60]:
            # AC options
            data.append([model, 110, 1, hz, 0])      # 1-phase 110V
            data.append([model, 220, 1, hz, 0])      # 1-phase 220V
            data.append([model, 380, 3, hz, 0])      # 3-phase 380V
            data.append([model, 440, 3, hz, 50])     # 3-phase 440V
            # DC option
            data.append([model, 24, "DC", hz, 50])   # DC 24V

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    widths = [22, 10, 8, 8, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = 'A2'


def setup_enclosure_options_db(ws, header_font, header_fill, border):
    """Setup DB_EnclosureOptions sheet (normalized)

    Each model can have multiple enclosure options.
    PriceAdder is added to (BasePrice + PowerOption PriceAdder).
    Model names must match DB_Models exactly.
    """
    headers = ["Model", "Enclosure", "PriceAdder"]

    data = []

    # ==================== NA Series Enclosure Options ====================
    na_all = ["NA006", "NA009", "NA015", "NA019", "NA028", "NA038", "NA050",
              "NA060", "NA080", "NA100", "NA120", "NA150", "NA200", "NA250", "NA300", "NA350"]

    for model in na_all:
        # Standard IP67 (waterproof)
        data.append([model, "IP67", 0])
        # Explosionproof option (larger models)
        if model in ["NA050", "NA060", "NA080", "NA100", "NA120", "NA150", "NA200", "NA250", "NA300", "NA350"]:
            data.append([model, "Exd", 300])

    # ==================== SA Series Enclosure Options ====================
    # ControlType: ONOFF, PCU, SCP (EXP는 Enclosure 옵션)
    # 방폭형 모델(SA05X, SA09X): Exd만 지원 (IP67 옵션 없음)
    sa_models = [
        ("SA003", ["ONOFF"]),
        ("SA005", ["ONOFF"]),
        ("SA005L", ["PCU", "SCP"]),
        ("SA05X", ["ONOFF"]),     # 방폭형 - Exd 기본
        ("SA009", ["ONOFF"]),
        ("SA009L", ["PCU", "SCP"]),
        ("SA09X", ["ONOFF"]),     # 방폭형 - Exd 기본
    ]

    for model, types in sa_models:
        is_exd_model = "X" in model  # SA05X, SA09X

        if is_exd_model:
            # 방폭형 모델: Exd만 지원 (가격 포함)
            data.append([model, "Exd", 0])
        else:
            # 일반 모델: IP67 기본, Exd 옵션
            data.append([model, "IP67", 0])
            data.append([model, "Exd", 200])

    # ==================== MS Series Enclosure Options ====================
    # MS01 enclosure options (same for both 3-phase and 1-phase)
    data.append(["MS01", "IP67", 0])
    data.append(["MS01", "Exd", 250])

    # ==================== MA Series Enclosure Options ====================
    ma_configs = [
        ("MA01", 200),
        ("MA02", 250),
        ("MA03", 350),
        ("MA04", 500),
        ("MA05", 700),
        ("MA06", 900),
    ]

    for model, exd_price in ma_configs:
        data.append([model, "IP67", 0])
        data.append([model, "Exd", exd_price])

    # ==================== SR Series Enclosure Options ====================
    # SR has normal (IP67) and X models (Exd)
    sr_models_list = ["SR05", "SR10", "SR20", "SR30", "SR50"]

    for model in sr_models_list:
        data.append([model, "IP67", 0])    # Standard waterproof
        data.append([model, "Exd", 300])   # Explosionproof option

    # ==================== NL Series Enclosure Options ====================
    nl_models_list = ["NL04", "NL06", "NL08", "NL10", "NL20", "NL25", "NL35"]

    for model in nl_models_list:
        data.append([model, "IP67", 0])
        data.append([model, "Exd", 250])  # Explosionproof option

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    widths = [22, 12, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = 'A2'


def setup_electrical_data_db(ws, header_font, header_fill, border):
    """Setup DB_ElectricalData sheet (normalized)

    Electrical characteristics depend on Model + Voltage/Phase/Freq combination.
    Used for Datasheet export.
    Model names must match DB_Models exactly.

    Columns (11 total):
    1. Model - Actuator model name
    2. Voltage - V
    3. Phase - 1, 3, or DC
    4. Freq - Hz (50 or 60)
    5. StartingCurrent_A - Starting current (placeholder if unknown)
    6. StartingPF - Starting power factor (placeholder if unknown)
    7. RatedCurrent_A - Rated load current
    8. AvgCurrent_A - Current at average load (placeholder if unknown)
    9. AvgPF - Power factor at average load (placeholder if unknown)
    10. AvgPower_kW - Motor power at average load
    11. MotorPoles - Number of motor poles

    Note: Some fields may be empty if catalog data is not available.
    These can be filled in manually or removed from datasheet if not needed.
    """
    headers = ["Model", "Voltage", "Phase", "Freq",
               "StartingCurrent_A", "StartingPF", "RatedCurrent_A",
               "AvgCurrent_A", "AvgPF", "AvgPower_kW", "MotorPoles"]

    data = []

    def add_row(model, voltage, phase, freq, rated_current, power_w, poles):
        """Helper to create 11-column row with placeholders for missing data.

        Placeholder fields (to be filled from catalog if available):
        - StartingCurrent_A: Typically 6-8x rated current
        - StartingPF: Typically 0.3-0.5
        - AvgCurrent_A: Typically less than rated
        - AvgPF: Typically 0.7-0.9
        """
        avg_power_kw = round(power_w / 1000, 3) if power_w else None
        return [model, voltage, phase, freq,
                None,           # StartingCurrent_A (placeholder)
                None,           # StartingPF (placeholder)
                rated_current,  # RatedCurrent_A
                None,           # AvgCurrent_A (placeholder)
                None,           # AvgPF (placeholder)
                avg_power_kw,   # AvgPower_kW
                poles]          # MotorPoles

    # ==================== NA Series Electrical Data ====================
    # From docs/noah_torque_tables.md - Current values at 60Hz
    # Columns: Model, DC12V, DC24V, AC110V, AC220V, AC230V, AC380V, AC440V, DC_W, AC_W
    na_electrical = [
        ("NA006", 7.3, 3.0, 0.85, 0.42, 0.42, 0.15, 0.15, 15, 15),
        ("NA009", 8.2, 4.3, 1.35, 0.55, 0.55, 0.19, 0.19, 25, 25),
        ("NA015", 10.2, 4.3, 1.75, 1.0, 1.0, 0.3, 0.33, 40, 40),
        ("NA019", 12, 5.0, 1.9, 1.05, 1.05, 0.31, 0.33, 40, 40),
        ("NA028", 14, 7.0, 2.0, 1.05, 1.05, 0.35, 0.32, 40, 40),
        ("NA038", 19.8, 13.5, 2.1, 1.3, 1.3, 0.4, 0.37, 120, 60),
        ("NA050", 24.5, 14.5, 3.8, 1.45, 1.45, 0.51, 0.5, 120, 90),
        ("NA060", 24.7, 15.0, 2.4, 1.5, 1.6, 0.47, 0.48, 120, 90),
        ("NA080", None, 16.0, 3.15, 1.9, 2.0, 0.62, 0.62, 120, 180),
        ("NA100", None, None, 3.5, 2.05, 2.1, 0.7, 0.7, None, 180),
        ("NA120", None, None, 4.0, 2.2, 2.4, 0.9, 0.75, None, 180),
        ("NA150", None, None, 2.4, 1.5, 1.6, 0.47, 0.48, None, 90),
        ("NA200", None, None, 3.15, 1.9, 2.0, 0.62, 0.62, None, 180),
        ("NA250", None, None, 3.5, 2.05, 2.1, 0.7, 0.7, None, 180),
        ("NA300", None, None, 4.0, 1.9, 2.4, 0.9, 0.75, None, 180),
        ("NA350", None, None, 4.0, 2.05, 2.4, 0.9, 0.75, None, 180),
    ]

    for model, dc12, dc24, ac110, ac220, ac230, ac380, ac440, dc_w, ac_w in na_electrical:
        for hz in [50, 60]:
            # AC options
            data.append(add_row(model, 380, 3, hz, ac380, ac_w, 4))
            data.append(add_row(model, 220, 1, hz, ac220, ac_w, 4))
            data.append(add_row(model, 110, 1, hz, ac110, ac_w, 4))
            data.append(add_row(model, 440, 3, hz, ac440, ac_w, 4))
            # DC options (if available)
            if dc24 is not None:
                data.append(add_row(model, 24, "DC", hz, dc24, dc_w, None))
            if dc12 is not None:
                data.append(add_row(model, 12, "DC", hz, dc12, dc_w, None))

    # ==================== SA Series Electrical Data ====================
    # From docs/noah_torque_tables.md
    # ControlType: ONOFF, PCU, SCP (EXP는 Enclosure 옵션)
    sa_electrical = [
        # Model, ControlType, AC24V, AC110V, AC220V, DC24V, Power_W
        ("SA003", "ONOFF", 1.0, 0.4, 0.16, 1.4, 25),
        ("SA005", "ONOFF", 1.8, 0.35, 0.23, 1.8, 6),
        ("SA005L", "PCU", 1.8, 0.35, 0.23, 1.8, 6),
        ("SA005L", "SCP", None, 0.8, 0.51, 2.25, 15),
        ("SA05X", "ONOFF", 1.8, 0.35, 0.23, 1.8, 6),   # 방폭형 - ControlType은 ONOFF
        ("SA009", "ONOFF", 2.1, 0.35, 0.25, 2.1, 6),
        ("SA009L", "PCU", 2.1, 0.35, 0.25, 2.1, 6),
        ("SA009L", "SCP", None, 0.8, 0.51, 2.25, 15),
        ("SA09X", "ONOFF", 2.1, 0.35, 0.25, 2.1, 6),   # 방폭형 - ControlType은 ONOFF
    ]

    for model, ctrl_type, ac24, ac110, ac220, dc24, power_w in sa_electrical:
        for hz in [50, 60]:
            data.append(add_row(model, 220, 1, hz, ac220, power_w, 4))
            data.append(add_row(model, 110, 1, hz, ac110, power_w, 4))
            if ac24 is not None:
                data.append(add_row(model, 24, 1, hz, ac24, power_w, 4))
            data.append(add_row(model, 24, "DC", hz, dc24, power_w, None))

    # ==================== MS Series Electrical Data ====================
    # Approximate data for MS01 (same Model, Phase distinguishes 3-phase vs 1-phase)
    for hz in [50, 60]:
        # MS01 3-phase
        data.append(add_row("MS01", 380, 3, hz, 0.8, 200, 4))
        data.append(add_row("MS01", 220, 3, hz, 1.4, 200, 4))
        data.append(add_row("MS01", 440, 3, hz, 0.7, 200, 4))
        # MS01 1-phase
        data.append(add_row("MS01", 220, 1, hz, 1.2, 150, 4))
        data.append(add_row("MS01", 110, 1, hz, 2.4, 150, 4))

    # ==================== MA Series Electrical Data ====================
    # Simplified - generate based on model (using average kW for current calculation)
    ma_configs = [
        ("MA01", 0.5),   # average kW
        ("MA02", 1.85),
        ("MA03", 4.7),
        ("MA04", 9.5),
        ("MA05", 16.5),
        ("MA06", 27),
    ]

    for model, avg_kw in ma_configs:
        power_w = int(avg_kw * 1000)
        # Current calculation: I = P / (V * sqrt(3) * PF) for 3-phase
        current_380 = round(power_w / (380 * 1.732 * 0.85), 2)
        current_440 = round(power_w / (440 * 1.732 * 0.85), 2)
        current_220 = round(power_w / (220 * 1.732 * 0.85), 2)

        # 50Hz options
        data.append(add_row(model, 380, 3, 50, current_380, power_w, 4))
        data.append(add_row(model, 220, 3, 50, current_220, power_w, 4))
        data.append(add_row(model, 440, 3, 50, current_440, power_w, 4))

        # 60Hz options
        data.append(add_row(model, 380, 3, 60, current_380, power_w, 4))
        data.append(add_row(model, 440, 3, 60, current_440, power_w, 4))

    # ==================== SR Series Electrical Data ====================
    # From user-provided SR specification table
    # Motor: AC 90W, DC 120W
    # Rated Current table provided per voltage
    # Model, 110V_50Hz, 110V_60Hz, 220V_50Hz, 220V_60Hz, 380V_50Hz, 380V_60Hz, 440V_50Hz, 440V_60Hz, DC_24V, AC_Power_W
    sr_electrical = [
        ("SR05", 2.8, 3.6, 1.6, 2.0, 0.24, 0.25, 0.28, 0.24, 12, 90),
        ("SR10", 2.8, 3.6, 1.6, 2.0, 0.4, 0.4, 0.5, 0.4, 12, 90),
        ("SR20", 2.8, 3.6, 1.7, 2.1, 0.38, 0.35, 0.5, 0.36, 13, 90),
        ("SR30", 2.8, 3.7, 1.6, 2.0, 0.35, 0.29, 0.62, 0.31, 15, 90),
        ("SR50", 2.8, 3.7, 1.6, 2.1, 0.37, 0.31, 0.62, 0.31, 15, 90),
    ]

    for model, i110_50, i110_60, i220_50, i220_60, i380_50, i380_60, i440_50, i440_60, dc24, power_w in sr_electrical:
        # 50Hz options
        data.append(add_row(model, 110, 1, 50, i110_50, power_w, 4))
        data.append(add_row(model, 120, 1, 50, i110_50, power_w, 4))  # 120V similar to 110V
        data.append(add_row(model, 220, 1, 50, i220_50, power_w, 4))
        data.append(add_row(model, 230, 1, 50, i220_50, power_w, 4))  # 230V similar to 220V
        data.append(add_row(model, 380, 3, 50, i380_50, power_w, 4))
        data.append(add_row(model, 440, 3, 50, i440_50, power_w, 4))
        data.append(add_row(model, 24, "DC", 50, dc24, 120, None))    # DC 120W

        # 60Hz options
        data.append(add_row(model, 110, 1, 60, i110_60, power_w, 4))
        data.append(add_row(model, 120, 1, 60, i110_60, power_w, 4))
        data.append(add_row(model, 220, 1, 60, i220_60, power_w, 4))
        data.append(add_row(model, 230, 1, 60, i220_60, power_w, 4))
        data.append(add_row(model, 380, 3, 60, i380_60, power_w, 4))
        data.append(add_row(model, 440, 3, 60, i440_60, power_w, 4))
        data.append(add_row(model, 24, "DC", 60, dc24, 120, None))

    # ==================== NL Series Electrical Data ====================
    # From user-provided NL specification table
    # Model, 110V_50Hz, 110V_60Hz, 220V_50Hz, 220V_60Hz, 380V_50Hz, 380V_60Hz, 440V_50Hz, 440V_60Hz, DC_24V, Motor_W
    nl_electrical = [
        ("NL04", 0.46, 0.43, 0.28, 0.27, 0.11, 0.09, 0.1, 0.09, 1.65, 15),
        ("NL06", 0.73, 0.72, 0.38, 0.37, 0.14, 0.12, 0.12, 0.11, 1.87, 25),
        ("NL08", 0.84, 0.81, 0.46, 0.43, 0.16, 0.14, 0.15, 0.13, 2.46, 25),
        ("NL10", 1.52, 1.5, 0.79, 0.75, 0.25, 0.24, 0.23, 0.21, 4.02, 40),
        ("NL20", 1.9, 1.95, 1.0, 1.5, 0.33, 0.33, 0.25, 0.32, 9.45, 60),
        ("NL25", 2.53, 3.25, 1.05, 1.45, 0.45, 0.45, 0.54, 0.45, 10.5, 90),
        ("NL35", 2.8, 3.6, 1.5, 2.0, 0.48, 0.49, 0.56, 0.51, 12.0, 90),
    ]

    for model, i110_50, i110_60, i220_50, i220_60, i380_50, i380_60, i440_50, i440_60, dc24, power_w in nl_electrical:
        # 50Hz options
        data.append(add_row(model, 110, 1, 50, i110_50, power_w, 4))
        data.append(add_row(model, 220, 1, 50, i220_50, power_w, 4))
        data.append(add_row(model, 380, 3, 50, i380_50, power_w, 4))
        data.append(add_row(model, 440, 3, 50, i440_50, power_w, 4))
        data.append(add_row(model, 24, "DC", 50, dc24, power_w, None))

        # 60Hz options
        data.append(add_row(model, 110, 1, 60, i110_60, power_w, 4))
        data.append(add_row(model, 220, 1, 60, i220_60, power_w, 4))
        data.append(add_row(model, 380, 3, 60, i380_60, power_w, 4))
        data.append(add_row(model, 440, 3, 60, i440_60, power_w, 4))
        data.append(add_row(model, 24, "DC", 60, dc24, power_w, None))

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    widths = [22, 10, 8, 8, 16, 12, 14, 14, 10, 14, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = 'A2'


def setup_gearboxes_db(ws, header_font, header_fill, border):
    """Setup DB_Gearboxes sheet with sample data"""

    headers = ["Model", "Ratio", "InputTorqueMax", "OutputTorqueMax",
               "Efficiency", "InputFlange", "OutputFlange", "MaxStemDim_mm",
               "Weight_kg", "Price"]

    # MaxStemDim_mm: Max valve stem diameter the gearbox output can handle
    # Sambo Gearbox Data (Representative Models)
    # Efficiency = Mechanical Advantage / Ratio
    # InputFlange = OutputFlange (assumed same)
    # Weight: 0 for now (to be added manually)
    # Price: placeholder (to be updated)

    data = [
        # Bevel Gear - SB-V Series (Part-turn / Multi-turn)
        # Model, Ratio, InputTorqueMax, OutputTorqueMax, Efficiency, InputFlange, OutputFlange, MaxStemDim, Weight, Price
        ["SB-VS10", 2.5, 92.6, 220, 0.96, "F10", "F10", 30, 0, 200],
        ["SB-VS10-1S", 5, 47.7, 220, 0.92, "F10", "F10", 30, 0, 250],
        ["SB-VS20", 3, 129.8, 370, 0.97, "F12", "F12", 40, 0, 280],
        ["SB-VS20-1S", 6, 66.9, 370, 0.92, "F12", "F12", 40, 0, 350],
        ["SB-V0", 3.25, 194.3, 600, 0.95, "F14", "F14", 46, 0, 400],
        ["SB-V0-1S", 6.5, 100.2, 600, 0.92, "F14", "F14", 46, 0, 500],
        ["SB-V0-1SD", 20.86, 32.1, 600, 0.90, "F14", "F14", 46, 0, 650],
        ["SB-V1", 3.5, 294.7, 980, 0.94, "F16", "F16", 55, 0, 550],
        ["SB-V1-1S", 7, 155.1, 980, 0.90, "F16", "F16", 55, 0, 700],
        ["SB-V1-1SD", 22.46, 48.5, 980, 0.90, "F16", "F16", 55, 0, 850],
        ["SB-V2", 4, 394.7, 1500, 0.95, "F14", "F14", 62, 0, 700],
        ["SB-V2-1S", 10.13, 160.6, 1500, 0.92, "F14", "F14", 62, 0, 900],
        ["SB-V2H", 4, 763.2, 2900, 0.95, "F25", "F25", 65, 0, 1200],
        ["SB-V3", 5, 526.3, 2500, 0.96, "F25", "F25", 72, 0, 900],
        ["SB-V3-1S", 12.67, 214.2, 2500, 0.92, "F25", "F25", 72, 0, 1100],
        ["SB-V3H", 5, 947.4, 4500, 0.96, "F30", "F30", 80, 0, 1500],
        ["SB-V4", 6, 912.3, 5200, 0.95, "F35", "F35", 98, 0, 1400],
        ["SB-V4-1S", 18, 313.5, 5200, 0.92, "F35", "F35", 98, 0, 1800],
        ["SB-V5", 6.56, 1252.5, 7800, 0.94, "F35", "F35", 110, 0, 2000],
        ["SB-V5-1S", 22.94, 368.9, 7800, 0.92, "F35", "F35", 110, 0, 2500],
        ["SB-V6", 7, 1954.9, 13000, 0.96, "F40", "F40", 130, 0, 3000],
        ["SB-V7", 7.56, 2452, 17600, 0.95, "F48", "F48", 150, 0, 4000],

        # Spur Gear - SB-SR Series (Multi-turn, high ratio)
        ["SB-SR50", 10, 40.2, 370, 0.92, "F12", "F12", 40, 0, 350],
        ["SB-SR50-2B", 20, 22.3, 370, 0.83, "F12", "F12", 40, 0, 450],
        ["SB-SR100", 12, 88.8, 980, 0.92, "F16", "F16", 55, 0, 500],
        ["SB-SR100-2B", 24, 49.3, 980, 0.83, "F16", "F16", 55, 0, 650],

        # Worm Gear - SBWG Series (Multi-turn, very high ratio, low efficiency)
        ["SBWG-BF", 32, 30.9, 360, 0.36, "F07", "F07", 20, 0, 300],
        ["SBWG-0", 36, 58.1, 780, 0.37, "F10", "F10", 28, 0, 400],
        ["SBWG-00", 40, 85.0, 1200, 0.35, "F12", "F12", 36, 0, 550],
        ["SBWG-00-1S", 80, 46.2, 1200, 0.33, "F12", "F12", 36, 0, 700],
        ["SBWG-00-1SD", 160, 25.1, 1200, 0.30, "F12", "F12", 36, 0, 900],
    ]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    # Write data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    widths = [14, 8, 15, 16, 12, 12, 12, 14, 12, 10]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = 'A2'


def setup_couplings_db(ws, header_font, header_fill, border):
    """Setup DB_Couplings sheet"""

    headers = ["CouplingType", "MinDimension_mm", "MaxDimension_mm"]

    data = [
        ["Thrust Base - Threaded", 20, 120],  # Multi-turn용
        ["Standard (Part-turn)", 0, 0],       # Part-turn용 (직접 플랜지 마운트)
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18


def setup_options_db(ws, header_font, header_fill, border):
    """Setup DB_Options sheet"""

    headers = ["Code", "Description", "Price"]

    data = [
        ["OPT-HTR", "Space Heater", 50],
        ["OPT-MOD", "Modulating Control", 200],
        ["OPT-EXD", "Explosionproof Upgrade", 300],
        ["OPT-POS", "Position Transmitter", 150],
        ["OPT-LMT", "Additional Limit Switch", 80],
        ["PAINT-EP", "Epoxy Coating", 100],
        ["PAINT-PU", "Polyurethane Coating", 150],
        ["PAINT-SPEC", "Special Coating", 250],
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10


def setup_datasheet_template(ws, header_font, border):
    """Setup Template_Datasheet sheet
    
    Rows 1-5: Reserved for logo/header (user can add logo here)
    Row 6+: Data starts
    """

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Header area (rows 1-5) for logo/company info
    HEADER_ROWS = 5
    
    # Set row heights for header area
    for r in range(1, HEADER_ROWS + 1):
        ws.row_dimensions[r].height = 20
    
    # Placeholder text for logo area
    ws['A1'] = "[Company Logo]"
    ws['A1'].font = Font(size=14, italic=True, color="808080")
    ws.merge_cells('A1:D3')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['A4'] = "ACTUATOR DATASHEET"
    ws['A4'].font = Font(bold=True, size=14)
    ws.merge_cells('A4:D4')
    ws['A4'].alignment = Alignment(horizontal='center')

    # Template structure based on the example datasheet
    # Data starts from row 6 (after 5 header rows)
    sections = [
        ("Item", "Units", "Line 1", "Line 2"),
        ("Line Number", "", "1", "2"),
        ("Tag Number", "", "", ""),
        ("Quantity", "Each", "1", "1"),
        ("", "", "", ""),
        ("Valve Requirements", "Units", "", ""),
        ("Type", "", "", ""),
        ("Size", "", "", ""),
        ("Class", "", "", ""),
        ("Torque", "Nm", "", ""),
        ("Thrust", "kN", "", ""),
        ("Coupling Type", "", "", ""),
        ("Coupling Dimension", "mm", "", ""),
        ("Turns", "", "", ""),
        ("Operating Time", "secs", "", ""),
        ("", "", "", ""),
        ("Equipment Offered", "Units", "", ""),
        ("Actuator", "", "", ""),
        ("Actuator Speed", "rpm", "", ""),
        ("Motor Power", "kW", "", ""),
        ("Secondary Gearbox", "", "", ""),
        ("Gearbox Ratio", "", "", ""),
        ("Output Flange", "", "", ""),
        ("Actuator Weight", "kg", "", ""),
        ("Gearbox Weight", "kg", "", ""),
        ("Combination Weight", "kg", "", ""),
        ("", "", "", ""),
        ("Actuator Performance", "Units", "", ""),
        ("Torque", "Nm", "", ""),
        ("Thrust", "kN", "", ""),
        ("Output Speed", "rpm", "", ""),
        ("Operating Time", "secs", "", ""),
        ("", "", "", ""),
        ("Safety Factors", "", "", ""),
        ("Requested - Torque", "", "", ""),
        ("Requested - Thrust", "", "", ""),
        ("Calculated - Torque", "", "", ""),
        ("Calculated - Thrust", "", "", ""),
        ("", "", "", ""),
        ("Electrical Data", "", "", ""),
        ("Voltage", "V", "", ""),
        ("Phase", "Ø", "", ""),
        ("Frequency", "Hz", "", ""),
        ("Starting current", "A", "", ""),
        ("Starting power factor", "", "", ""),
        ("Rated load current", "A", "", ""),
        ("Current at average load", "A", "", ""),
        ("Power factor at average load", "", "", ""),
        ("Motor power at average load", "kW", "", ""),
        ("Number of poles of motor", "", "", ""),
    ]

    # Start from row 6 (after header area)
    START_ROW = HEADER_ROWS + 1
    
    for row_idx, row_data in enumerate(sections, START_ROW):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 1 and value and value not in ["", "Item"]:
                cell.font = Font(bold=True) if "Requirements" in value or "Offered" in value or "Performance" in value or "Factors" in value or "Data" in value else None
            cell.border = border


if __name__ == "__main__":
    wb = create_workbook()
    wb.save("NoahSizing.xlsx")
    print("NoahSizing.xlsx created successfully!")
    print("\nNext steps:")
    print("1. Open NoahSizing.xlsx in Excel")
    print("2. Save as NoahSizing.xlsm (Excel Macro-Enabled Workbook)")
    print("3. Import VBA modules from the 'vba' folder")
